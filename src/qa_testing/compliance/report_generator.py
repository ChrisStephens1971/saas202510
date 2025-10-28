"""
Compliance Report Generator for Accounting System

Generates standard financial reports required for compliance and auditing:
- General Ledger (GL): Complete transaction history with running balances
- Trial Balance (TB): Summary of account balances verifying debits = credits
- Export formats: PDF, Excel, CSV

All reports support multi-tenant isolation and date range filtering.
"""

from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import Optional
from uuid import UUID

from pydantic import BaseModel, Field

from ..models import LedgerEntry, Fund


class ReportFormat(str, Enum):
    """Supported report export formats."""
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"


class GeneralLedgerEntry(BaseModel):
    """Single entry in a General Ledger report."""
    entry_date: date
    entry_id: UUID
    fund_name: str
    description: str
    debit_amount: Optional[Decimal] = None
    credit_amount: Optional[Decimal] = None
    running_balance: Decimal
    transaction_id: Optional[UUID] = None


class GeneralLedgerReport(BaseModel):
    """General Ledger report for a date range."""
    tenant_id: UUID
    report_date: datetime
    start_date: date
    end_date: date
    fund_id: Optional[UUID] = None  # None = all funds
    fund_name: Optional[str] = None

    # Entries
    entries: list[GeneralLedgerEntry] = Field(default_factory=list)

    # Summary
    total_debits: Decimal = Decimal("0.00")
    total_credits: Decimal = Decimal("0.00")
    opening_balance: Decimal = Decimal("0.00")
    closing_balance: Decimal = Decimal("0.00")
    entry_count: int = 0

    # Validation
    is_balanced: bool = True
    balance_difference: Decimal = Decimal("0.00")


class TrialBalanceAccount(BaseModel):
    """Single account in a Trial Balance report."""
    fund_id: UUID
    fund_name: str
    debit_balance: Decimal = Decimal("0.00")
    credit_balance: Decimal = Decimal("0.00")
    net_balance: Decimal = Decimal("0.00")
    entry_count: int = 0


class TrialBalanceReport(BaseModel):
    """Trial Balance report at a point in time."""
    tenant_id: UUID
    report_date: datetime
    as_of_date: date

    # Accounts
    accounts: list[TrialBalanceAccount] = Field(default_factory=list)

    # Summary
    total_debits: Decimal = Decimal("0.00")
    total_credits: Decimal = Decimal("0.00")
    account_count: int = 0

    # Validation
    is_balanced: bool = True
    balance_difference: Decimal = Decimal("0.00")


class ComplianceReportGenerator:
    """
    Generate compliance reports for accounting system.

    Reports:
    - General Ledger: Complete transaction history
    - Trial Balance: Account balance summary

    Export formats: PDF, Excel, CSV
    """

    @staticmethod
    def generate_general_ledger(
        tenant_id: UUID,
        ledger_entries: list[LedgerEntry],
        funds: list[Fund],
        start_date: date,
        end_date: date,
        fund_id: Optional[UUID] = None,
    ) -> GeneralLedgerReport:
        """
        Generate General Ledger report.

        Args:
            tenant_id: Tenant for report
            ledger_entries: All ledger entries to include
            funds: All funds for lookup
            start_date: Report start date
            end_date: Report end date (inclusive)
            fund_id: Filter by specific fund (None = all funds)

        Returns:
            GeneralLedgerReport with entries and summary
        """
        # Filter by tenant and date range
        filtered_entries = [
            entry for entry in ledger_entries
            if entry.tenant_id == tenant_id
            and start_date <= entry.entry_date <= end_date
        ]

        # Filter by fund if specified
        if fund_id:
            filtered_entries = [e for e in filtered_entries if e.fund_id == fund_id]

        # Sort by date, then created_at
        filtered_entries.sort(key=lambda e: (e.entry_date, e.created_at))

        # Build fund lookup
        fund_lookup = {f.id: f for f in funds}

        # Calculate opening balance (entries before start_date)
        opening_balance = Decimal("0.00")
        for entry in ledger_entries:
            if entry.tenant_id == tenant_id and entry.entry_date < start_date:
                if fund_id is None or entry.fund_id == fund_id:
                    if entry.is_debit:
                        opening_balance += entry.amount
                    else:
                        opening_balance -= entry.amount

        # Build GL entries with running balance
        gl_entries: list[GeneralLedgerEntry] = []
        running_balance = opening_balance
        total_debits = Decimal("0.00")
        total_credits = Decimal("0.00")

        for entry in filtered_entries:
            # Update running balance
            if entry.is_debit:
                running_balance += entry.amount
                total_debits += entry.amount
                debit_amount = entry.amount
                credit_amount = None
            else:
                running_balance -= entry.amount
                total_credits += entry.amount
                debit_amount = None
                credit_amount = entry.amount

            # Get fund name
            fund = fund_lookup.get(entry.fund_id)
            fund_name = fund.name if fund else "Unknown Fund"

            # Create GL entry
            gl_entry = GeneralLedgerEntry(
                entry_date=entry.entry_date,
                entry_id=entry.id,
                fund_name=fund_name,
                description=entry.description,
                debit_amount=debit_amount,
                credit_amount=credit_amount,
                running_balance=running_balance,
                transaction_id=entry.transaction_id,
            )
            gl_entries.append(gl_entry)

        # Determine fund name for report
        report_fund_name = None
        if fund_id:
            fund = fund_lookup.get(fund_id)
            report_fund_name = fund.name if fund else "Unknown Fund"

        # Calculate balance difference
        balance_diff = total_debits - total_credits
        is_balanced = abs(balance_diff) < Decimal("0.01")  # Allow 1 cent rounding

        return GeneralLedgerReport(
            tenant_id=tenant_id,
            report_date=datetime.now(),
            start_date=start_date,
            end_date=end_date,
            fund_id=fund_id,
            fund_name=report_fund_name,
            entries=gl_entries,
            total_debits=total_debits,
            total_credits=total_credits,
            opening_balance=opening_balance,
            closing_balance=running_balance,
            entry_count=len(gl_entries),
            is_balanced=is_balanced,
            balance_difference=balance_diff,
        )

    @staticmethod
    def generate_trial_balance(
        tenant_id: UUID,
        ledger_entries: list[LedgerEntry],
        funds: list[Fund],
        as_of_date: date,
    ) -> TrialBalanceReport:
        """
        Generate Trial Balance report.

        Args:
            tenant_id: Tenant for report
            ledger_entries: All ledger entries up to as_of_date
            funds: All funds for lookup
            as_of_date: Report date (inclusive)

        Returns:
            TrialBalanceReport with account balances and summary
        """
        # Filter by tenant and date
        filtered_entries = [
            entry for entry in ledger_entries
            if entry.tenant_id == tenant_id
            and entry.entry_date <= as_of_date
        ]

        # Build fund lookup
        fund_lookup = {f.id: f for f in funds}

        # Calculate balances by fund
        fund_balances: dict[UUID, dict] = {}

        for entry in filtered_entries:
            if entry.fund_id not in fund_balances:
                fund_balances[entry.fund_id] = {
                    "debit": Decimal("0.00"),
                    "credit": Decimal("0.00"),
                    "count": 0,
                }

            if entry.is_debit:
                fund_balances[entry.fund_id]["debit"] += entry.amount
            else:
                fund_balances[entry.fund_id]["credit"] += entry.amount

            fund_balances[entry.fund_id]["count"] += 1

        # Build TB accounts
        tb_accounts: list[TrialBalanceAccount] = []
        total_debits = Decimal("0.00")
        total_credits = Decimal("0.00")

        for fund_id, balances in fund_balances.items():
            fund = fund_lookup.get(fund_id)
            fund_name = fund.name if fund else "Unknown Fund"

            debit_bal = balances["debit"]
            credit_bal = balances["credit"]
            net_bal = debit_bal - credit_bal

            total_debits += debit_bal
            total_credits += credit_bal

            tb_account = TrialBalanceAccount(
                fund_id=fund_id,
                fund_name=fund_name,
                debit_balance=debit_bal,
                credit_balance=credit_bal,
                net_balance=net_bal,
                entry_count=balances["count"],
            )
            tb_accounts.append(tb_account)

        # Sort by fund name
        tb_accounts.sort(key=lambda a: a.fund_name)

        # Calculate balance difference
        balance_diff = total_debits - total_credits
        is_balanced = abs(balance_diff) < Decimal("0.01")  # Allow 1 cent rounding

        return TrialBalanceReport(
            tenant_id=tenant_id,
            report_date=datetime.now(),
            as_of_date=as_of_date,
            accounts=tb_accounts,
            total_debits=total_debits,
            total_credits=total_credits,
            account_count=len(tb_accounts),
            is_balanced=is_balanced,
            balance_difference=balance_diff,
        )

    @staticmethod
    def export_general_ledger_pdf(
        report: GeneralLedgerReport,
        output_path: Path,
    ) -> None:
        """
        Export General Ledger report to PDF.

        Args:
            report: GeneralLedgerReport to export
            output_path: Path to save PDF file
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        # Create PDF
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch,
        )

        # Build content
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph("<b>General Ledger Report</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))

        # Report info
        info_text = f"""
        <b>Period:</b> {report.start_date} to {report.end_date}<br/>
        <b>Fund:</b> {report.fund_name or "All Funds"}<br/>
        <b>Tenant ID:</b> {report.tenant_id}<br/>
        <b>Generated:</b> {report.report_date.strftime("%Y-%m-%d %H:%M:%S")}<br/>
        """
        info = Paragraph(info_text, styles['Normal'])
        elements.append(info)
        elements.append(Spacer(1, 0.3*inch))

        # Opening balance
        opening_text = f"<b>Opening Balance:</b> ${report.opening_balance:,.2f}"
        elements.append(Paragraph(opening_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))

        # Table header
        data = [[
            'Date', 'Fund', 'Description', 'Debit', 'Credit', 'Balance'
        ]]

        # Table rows
        for entry in report.entries:
            debit_str = f"${entry.debit_amount:,.2f}" if entry.debit_amount else ""
            credit_str = f"${entry.credit_amount:,.2f}" if entry.credit_amount else ""

            data.append([
                entry.entry_date.strftime("%Y-%m-%d"),
                entry.fund_name,
                entry.description[:40],  # Truncate long descriptions
                debit_str,
                credit_str,
                f"${entry.running_balance:,.2f}",
            ])

        # Summary row
        data.append([
            '', '', '<b>TOTAL</b>',
            f"<b>${report.total_debits:,.2f}</b>",
            f"<b>${report.total_credits:,.2f}</b>",
            f"<b>${report.closing_balance:,.2f}</b>",
        ])

        # Create table
        table = Table(data, colWidths=[1*inch, 1.5*inch, 3*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data rows
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('ALIGN', (3, 1), (5, -2), 'RIGHT'),  # Numbers right-aligned
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),

            # Summary row
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN', (3, -1), (5, -1), 'RIGHT'),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))

        # Balance status
        if report.is_balanced:
            status_text = "<b>Status:</b> <font color='green'>Balanced</font>"
        else:
            status_text = f"<b>Status:</b> <font color='red'>Out of Balance by ${abs(report.balance_difference):,.2f}</font>"

        elements.append(Paragraph(status_text, styles['Normal']))

        # Build PDF
        doc.build(elements)

    @staticmethod
    def export_general_ledger_excel(
        report: GeneralLedgerReport,
        output_path: Path,
    ) -> None:
        """
        Export General Ledger report to Excel.

        Args:
            report: GeneralLedgerReport to export
            output_path: Path to save Excel file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "General Ledger"

        # Title
        ws['A1'] = "General Ledger Report"
        ws['A1'].font = Font(size=16, bold=True)

        # Report info
        ws['A3'] = "Period:"
        ws['B3'] = f"{report.start_date} to {report.end_date}"
        ws['A4'] = "Fund:"
        ws['B4'] = report.fund_name or "All Funds"
        ws['A5'] = "Tenant ID:"
        ws['B5'] = str(report.tenant_id)
        ws['A6'] = "Generated:"
        ws['B6'] = report.report_date.strftime("%Y-%m-%d %H:%M:%S")

        # Opening balance
        ws['A8'] = "Opening Balance:"
        ws['B8'] = float(report.opening_balance)
        ws['B8'].number_format = '$#,##0.00'
        ws['B8'].font = Font(bold=True)

        # Header row
        headers = ['Date', 'Fund', 'Description', 'Debit', 'Credit', 'Balance']
        header_row = 10
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        row = header_row + 1
        for entry in report.entries:
            ws.cell(row=row, column=1).value = entry.entry_date
            ws.cell(row=row, column=2).value = entry.fund_name
            ws.cell(row=row, column=3).value = entry.description

            if entry.debit_amount:
                ws.cell(row=row, column=4).value = float(entry.debit_amount)
                ws.cell(row=row, column=4).number_format = '$#,##0.00'

            if entry.credit_amount:
                ws.cell(row=row, column=5).value = float(entry.credit_amount)
                ws.cell(row=row, column=5).number_format = '$#,##0.00'

            ws.cell(row=row, column=6).value = float(entry.running_balance)
            ws.cell(row=row, column=6).number_format = '$#,##0.00'

            row += 1

        # Summary row
        summary_row = row
        ws.cell(row=summary_row, column=3).value = "TOTAL"
        ws.cell(row=summary_row, column=3).font = Font(bold=True)

        ws.cell(row=summary_row, column=4).value = float(report.total_debits)
        ws.cell(row=summary_row, column=4).number_format = '$#,##0.00'
        ws.cell(row=summary_row, column=4).font = Font(bold=True)

        ws.cell(row=summary_row, column=5).value = float(report.total_credits)
        ws.cell(row=summary_row, column=5).number_format = '$#,##0.00'
        ws.cell(row=summary_row, column=5).font = Font(bold=True)

        ws.cell(row=summary_row, column=6).value = float(report.closing_balance)
        ws.cell(row=summary_row, column=6).number_format = '$#,##0.00'
        ws.cell(row=summary_row, column=6).font = Font(bold=True)

        # Status
        ws.cell(row=summary_row + 2, column=1).value = "Status:"
        ws.cell(row=summary_row + 2, column=1).font = Font(bold=True)

        if report.is_balanced:
            ws.cell(row=summary_row + 2, column=2).value = "Balanced"
            ws.cell(row=summary_row + 2, column=2).font = Font(color="00FF00", bold=True)
        else:
            ws.cell(row=summary_row + 2, column=2).value = f"Out of Balance by ${abs(report.balance_difference):,.2f}"
            ws.cell(row=summary_row + 2, column=2).font = Font(color="FF0000", bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        # Save
        wb.save(output_path)

    @staticmethod
    def export_trial_balance_pdf(
        report: TrialBalanceReport,
        output_path: Path,
    ) -> None:
        """
        Export Trial Balance report to PDF.

        Args:
            report: TrialBalanceReport to export
            output_path: Path to save PDF file
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        # Create PDF
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=letter,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch,
        )

        # Build content
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph("<b>Trial Balance Report</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))

        # Report info
        info_text = f"""
        <b>As of Date:</b> {report.as_of_date}<br/>
        <b>Tenant ID:</b> {report.tenant_id}<br/>
        <b>Generated:</b> {report.report_date.strftime("%Y-%m-%d %H:%M:%S")}<br/>
        """
        info = Paragraph(info_text, styles['Normal'])
        elements.append(info)
        elements.append(Spacer(1, 0.3*inch))

        # Table header
        data = [[
            'Fund', 'Debit Balance', 'Credit Balance', 'Net Balance'
        ]]

        # Table rows
        for account in report.accounts:
            data.append([
                account.fund_name,
                f"${account.debit_balance:,.2f}",
                f"${account.credit_balance:,.2f}",
                f"${account.net_balance:,.2f}",
            ])

        # Summary row
        data.append([
            '<b>TOTAL</b>',
            f"<b>${report.total_debits:,.2f}</b>",
            f"<b>${report.total_credits:,.2f}</b>",
            '',
        ])

        # Create table
        table = Table(data, colWidths=[3.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data rows
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 10),
            ('ALIGN', (1, 1), (3, -2), 'RIGHT'),  # Numbers right-aligned
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),

            # Summary row
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, -1), (2, -1), 'RIGHT'),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))

        # Balance status
        if report.is_balanced:
            status_text = "<b>Status:</b> <font color='green'>Balanced (Debits = Credits)</font>"
        else:
            status_text = f"<b>Status:</b> <font color='red'>Out of Balance by ${abs(report.balance_difference):,.2f}</font>"

        elements.append(Paragraph(status_text, styles['Normal']))

        # Build PDF
        doc.build(elements)

    @staticmethod
    def export_trial_balance_excel(
        report: TrialBalanceReport,
        output_path: Path,
    ) -> None:
        """
        Export Trial Balance report to Excel.

        Args:
            report: TrialBalanceReport to export
            output_path: Path to save Excel file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Trial Balance"

        # Title
        ws['A1'] = "Trial Balance Report"
        ws['A1'].font = Font(size=16, bold=True)

        # Report info
        ws['A3'] = "As of Date:"
        ws['B3'] = str(report.as_of_date)
        ws['A4'] = "Tenant ID:"
        ws['B4'] = str(report.tenant_id)
        ws['A5'] = "Generated:"
        ws['B5'] = report.report_date.strftime("%Y-%m-%d %H:%M:%S")

        # Header row
        headers = ['Fund', 'Debit Balance', 'Credit Balance', 'Net Balance']
        header_row = 7
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        row = header_row + 1
        for account in report.accounts:
            ws.cell(row=row, column=1).value = account.fund_name

            ws.cell(row=row, column=2).value = float(account.debit_balance)
            ws.cell(row=row, column=2).number_format = '$#,##0.00'

            ws.cell(row=row, column=3).value = float(account.credit_balance)
            ws.cell(row=row, column=3).number_format = '$#,##0.00'

            ws.cell(row=row, column=4).value = float(account.net_balance)
            ws.cell(row=row, column=4).number_format = '$#,##0.00'

            row += 1

        # Summary row
        summary_row = row
        ws.cell(row=summary_row, column=1).value = "TOTAL"
        ws.cell(row=summary_row, column=1).font = Font(bold=True)

        ws.cell(row=summary_row, column=2).value = float(report.total_debits)
        ws.cell(row=summary_row, column=2).number_format = '$#,##0.00'
        ws.cell(row=summary_row, column=2).font = Font(bold=True)

        ws.cell(row=summary_row, column=3).value = float(report.total_credits)
        ws.cell(row=summary_row, column=3).number_format = '$#,##0.00'
        ws.cell(row=summary_row, column=3).font = Font(bold=True)

        # Status
        ws.cell(row=summary_row + 2, column=1).value = "Status:"
        ws.cell(row=summary_row + 2, column=1).font = Font(bold=True)

        if report.is_balanced:
            ws.cell(row=summary_row + 2, column=2).value = "Balanced (Debits = Credits)"
            ws.cell(row=summary_row + 2, column=2).font = Font(color="00FF00", bold=True)
        else:
            ws.cell(row=summary_row + 2, column=2).value = f"Out of Balance by ${abs(report.balance_difference):,.2f}"
            ws.cell(row=summary_row + 2, column=2).font = Font(color="FF0000", bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

        # Save
        wb.save(output_path)
