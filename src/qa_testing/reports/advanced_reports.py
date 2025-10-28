"""
Advanced Report Generation with Templates and Customization

Provides enhanced financial statement exports with:
- Custom report templates
- Flexible filtering (date, fund, property)
- Column selection and customization
- Multiple output formats (PDF, Excel, CSV)
- Report scheduling support
"""

from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from typing import Optional, Any
from uuid import UUID, uuid4
import csv
from io import StringIO, BytesIO

from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from qa_testing.models import Transaction, LedgerEntry, Member, Fund


class ReportFormat(str, Enum):
    """Supported report output formats."""
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"
    JSON = "json"


class ColumnType(str, Enum):
    """Types of columns available in reports."""
    DATE = "date"
    TEXT = "text"
    AMOUNT = "amount"
    ACCOUNT = "account"
    REFERENCE = "reference"
    DESCRIPTION = "description"
    BALANCE = "balance"


class ReportColumn(BaseModel):
    """Definition of a report column."""
    key: str = Field(..., description="Data key for this column")
    label: str = Field(..., description="Display label for column header")
    column_type: ColumnType = Field(..., description="Type of data in column")
    width: Optional[int] = Field(None, description="Column width (format-specific)")
    align: str = Field("left", description="Text alignment (left, center, right)")
    format_string: Optional[str] = Field(None, description="Format string for values")


class ReportFilter(BaseModel):
    """Filtering criteria for reports."""
    tenant_id: UUID = Field(..., description="Tenant ID for isolation")
    start_date: Optional[date] = Field(None, description="Start date for filtering")
    end_date: Optional[date] = Field(None, description="End date for filtering")
    fund_ids: Optional[list[UUID]] = Field(None, description="Filter by specific funds")
    property_ids: Optional[list[UUID]] = Field(None, description="Filter by specific properties")
    member_ids: Optional[list[UUID]] = Field(None, description="Filter by specific members")
    account_codes: Optional[list[str]] = Field(None, description="Filter by account codes")
    min_amount: Optional[Decimal] = Field(None, description="Minimum transaction amount")
    max_amount: Optional[Decimal] = Field(None, description="Maximum transaction amount")
    include_voided: bool = Field(False, description="Include voided transactions")


class ReportTemplate(BaseModel):
    """Template definition for customizable reports."""
    template_id: UUID = Field(default_factory=uuid4, description="Unique template ID")
    name: str = Field(..., description="Template name")
    description: Optional[str] = Field(None, description="Template description")
    report_type: str = Field(..., description="Type of report (balance_sheet, income_statement, etc.)")
    columns: list[ReportColumn] = Field(..., description="Columns to include in report")
    default_format: ReportFormat = Field(ReportFormat.PDF, description="Default output format")
    page_size: str = Field("letter", description="Page size for PDF (letter, A4)")
    orientation: str = Field("portrait", description="Page orientation (portrait, landscape)")
    show_header: bool = Field(True, description="Show report header")
    show_footer: bool = Field(True, description="Show report footer")
    show_summary: bool = Field(True, description="Show summary totals")
    group_by: Optional[str] = Field(None, description="Field to group results by")
    sort_by: Optional[str] = Field(None, description="Field to sort results by")
    sort_order: str = Field("asc", description="Sort order (asc, desc)")


class ReportData(BaseModel):
    """Generated report data."""
    report_id: UUID = Field(default_factory=uuid4)
    template_id: UUID
    generated_at: datetime = Field(default_factory=datetime.now)
    filter_applied: ReportFilter
    row_count: int
    rows: list[dict[str, Any]]
    summary: Optional[dict[str, Any]] = None
    metadata: dict[str, Any] = Field(default_factory=dict)


class AdvancedReportGenerator:
    """
    Advanced report generator with templating and customization.

    Features:
    - Custom report templates
    - Flexible filtering
    - Multiple output formats
    - Column customization
    """

    @staticmethod
    def create_balance_sheet_template(
        columns: Optional[list[str]] = None
    ) -> ReportTemplate:
        """
        Create a standard balance sheet template.

        Args:
            columns: Optional list of column keys to include

        Returns:
            ReportTemplate for balance sheet
        """
        all_columns = [
            ReportColumn(
                key="account_code",
                label="Account",
                column_type=ColumnType.ACCOUNT,
                width=15,
                align="left"
            ),
            ReportColumn(
                key="account_name",
                label="Description",
                column_type=ColumnType.DESCRIPTION,
                width=30,
                align="left"
            ),
            ReportColumn(
                key="debit_balance",
                label="Debit",
                column_type=ColumnType.AMOUNT,
                width=15,
                align="right",
                format_string="${:,.2f}"
            ),
            ReportColumn(
                key="credit_balance",
                label="Credit",
                column_type=ColumnType.AMOUNT,
                width=15,
                align="right",
                format_string="${:,.2f}"
            ),
            ReportColumn(
                key="net_balance",
                label="Balance",
                column_type=ColumnType.BALANCE,
                width=15,
                align="right",
                format_string="${:,.2f}"
            ),
        ]

        # Filter columns if specific ones requested
        if columns:
            all_columns = [col for col in all_columns if col.key in columns]

        return ReportTemplate(
            name="Balance Sheet",
            description="Standard balance sheet report",
            report_type="balance_sheet",
            columns=all_columns,
            default_format=ReportFormat.PDF,
            show_summary=True,
            sort_by="account_code",
            sort_order="asc"
        )

    @staticmethod
    def create_income_statement_template(
        columns: Optional[list[str]] = None
    ) -> ReportTemplate:
        """
        Create a standard income statement template.

        Args:
            columns: Optional list of column keys to include

        Returns:
            ReportTemplate for income statement
        """
        all_columns = [
            ReportColumn(
                key="account_code",
                label="Account",
                column_type=ColumnType.ACCOUNT,
                width=15,
                align="left"
            ),
            ReportColumn(
                key="account_name",
                label="Description",
                column_type=ColumnType.DESCRIPTION,
                width=30,
                align="left"
            ),
            ReportColumn(
                key="amount",
                label="Amount",
                column_type=ColumnType.AMOUNT,
                width=15,
                align="right",
                format_string="${:,.2f}"
            ),
            ReportColumn(
                key="percentage",
                label="% of Total",
                column_type=ColumnType.TEXT,
                width=10,
                align="right",
                format_string="{:.1f}%"
            ),
        ]

        if columns:
            all_columns = [col for col in all_columns if col.key in columns]

        return ReportTemplate(
            name="Income Statement",
            description="Standard income statement report",
            report_type="income_statement",
            columns=all_columns,
            default_format=ReportFormat.PDF,
            show_summary=True,
            group_by="account_type",
            sort_by="account_code",
            sort_order="asc"
        )

    @staticmethod
    def create_transaction_detail_template(
        columns: Optional[list[str]] = None
    ) -> ReportTemplate:
        """
        Create a transaction detail report template.

        Args:
            columns: Optional list of column keys to include

        Returns:
            ReportTemplate for transaction details
        """
        all_columns = [
            ReportColumn(
                key="transaction_date",
                label="Date",
                column_type=ColumnType.DATE,
                width=12,
                align="left"
            ),
            ReportColumn(
                key="reference",
                label="Reference",
                column_type=ColumnType.REFERENCE,
                width=15,
                align="left"
            ),
            ReportColumn(
                key="description",
                label="Description",
                column_type=ColumnType.DESCRIPTION,
                width=30,
                align="left"
            ),
            ReportColumn(
                key="account",
                label="Account",
                column_type=ColumnType.ACCOUNT,
                width=15,
                align="left"
            ),
            ReportColumn(
                key="debit",
                label="Debit",
                column_type=ColumnType.AMOUNT,
                width=15,
                align="right",
                format_string="${:,.2f}"
            ),
            ReportColumn(
                key="credit",
                label="Credit",
                column_type=ColumnType.AMOUNT,
                width=15,
                align="right",
                format_string="${:,.2f}"
            ),
        ]

        if columns:
            all_columns = [col for col in all_columns if col.key in columns]

        return ReportTemplate(
            name="Transaction Detail",
            description="Detailed transaction listing",
            report_type="transaction_detail",
            columns=all_columns,
            default_format=ReportFormat.EXCEL,
            show_summary=True,
            sort_by="transaction_date",
            sort_order="asc"
        )

    @staticmethod
    def generate_report(
        template: ReportTemplate,
        data: list[dict[str, Any]],
        filters: ReportFilter,
        calculate_summary: bool = True
    ) -> ReportData:
        """
        Generate report data from template and source data.

        Args:
            template: Report template to use
            data: Source data for report
            filters: Filters applied to data
            calculate_summary: Whether to calculate summary totals

        Returns:
            ReportData with processed results
        """
        # Filter data based on filters
        filtered_data = AdvancedReportGenerator._apply_filters(data, filters)

        # Sort data
        if template.sort_by:
            reverse = template.sort_order == "desc"
            filtered_data = sorted(
                filtered_data,
                key=lambda x: x.get(template.sort_by, ""),
                reverse=reverse
            )

        # Calculate summary if requested
        summary = None
        if calculate_summary and template.show_summary:
            summary = AdvancedReportGenerator._calculate_summary(
                filtered_data,
                template
            )

        return ReportData(
            template_id=template.template_id,
            filter_applied=filters,
            row_count=len(filtered_data),
            rows=filtered_data,
            summary=summary,
            metadata={
                "template_name": template.name,
                "report_type": template.report_type
            }
        )

    @staticmethod
    def _apply_filters(
        data: list[dict[str, Any]],
        filters: ReportFilter
    ) -> list[dict[str, Any]]:
        """Apply filters to data."""
        filtered = data

        # Filter by date range
        if filters.start_date:
            filtered = [
                row for row in filtered
                if "transaction_date" in row and row["transaction_date"] >= filters.start_date
            ]
        if filters.end_date:
            filtered = [
                row for row in filtered
                if "transaction_date" in row and row["transaction_date"] <= filters.end_date
            ]

        # Filter by funds
        if filters.fund_ids:
            filtered = [
                row for row in filtered
                if "fund_id" in row and row["fund_id"] in filters.fund_ids
            ]

        # Filter by properties
        if filters.property_ids:
            filtered = [
                row for row in filtered
                if "property_id" in row and row["property_id"] in filters.property_ids
            ]

        # Filter by members
        if filters.member_ids:
            filtered = [
                row for row in filtered
                if "member_id" in row and row["member_id"] in filters.member_ids
            ]

        # Filter by amount range
        if filters.min_amount is not None:
            filtered = [
                row for row in filtered
                if "amount" in row and Decimal(str(row["amount"])) >= filters.min_amount
            ]
        if filters.max_amount is not None:
            filtered = [
                row for row in filtered
                if "amount" in row and Decimal(str(row["amount"])) <= filters.max_amount
            ]

        return filtered

    @staticmethod
    def _calculate_summary(
        data: list[dict[str, Any]],
        template: ReportTemplate
    ) -> dict[str, Any]:
        """Calculate summary totals for report."""
        summary: dict[str, Any] = {
            "total_rows": len(data)
        }

        # Sum amount columns
        amount_columns = [
            col.key for col in template.columns
            if col.column_type == ColumnType.AMOUNT
        ]

        for col_key in amount_columns:
            total = Decimal("0")
            for row in data:
                if col_key in row and row[col_key] is not None:
                    total += Decimal(str(row[col_key]))
            summary[f"total_{col_key}"] = float(total)

        return summary

    @staticmethod
    def export_to_csv(report_data: ReportData, template: ReportTemplate) -> str:
        """
        Export report to CSV format.

        Args:
            report_data: Generated report data
            template: Template used for report

        Returns:
            CSV string
        """
        output = StringIO()
        writer = csv.writer(output)

        # Write header row
        headers = [col.label for col in template.columns]
        writer.writerow(headers)

        # Write data rows
        for row in report_data.rows:
            csv_row = []
            for col in template.columns:
                value = row.get(col.key, "")
                # Format value if format string provided
                if col.format_string and value:
                    try:
                        value = col.format_string.format(value)
                    except:
                        pass
                csv_row.append(value)
            writer.writerow(csv_row)

        # Write summary if available
        if report_data.summary and template.show_summary:
            writer.writerow([])  # Empty row
            writer.writerow(["SUMMARY"])
            for key, value in report_data.summary.items():
                writer.writerow([key, value])

        return output.getvalue()

    @staticmethod
    def export_to_excel(
        report_data: ReportData,
        template: ReportTemplate,
        filepath: str
    ) -> None:
        """
        Export report to Excel format.

        Args:
            report_data: Generated report data
            template: Template used for report
            filepath: Output file path
        """
        wb = Workbook()
        ws = wb.active
        ws.title = template.name[:31]  # Excel sheet name limit

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col_idx, col in enumerate(template.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col.label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

            # Set column width
            if col.width:
                ws.column_dimensions[chr(64 + col_idx)].width = col.width

        # Write data rows
        for row_idx, data_row in enumerate(report_data.rows, start=2):
            for col_idx, col in enumerate(template.columns, start=1):
                value = data_row.get(col.key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

                # Apply alignment
                if col.align == "right":
                    cell.alignment = Alignment(horizontal="right")
                elif col.align == "center":
                    cell.alignment = Alignment(horizontal="center")

                # Apply number format for amounts
                if col.column_type == ColumnType.AMOUNT:
                    cell.number_format = '$#,##0.00'

        # Write summary if available
        if report_data.summary and template.show_summary:
            summary_row = len(report_data.rows) + 3
            ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True)

            for idx, (key, value) in enumerate(report_data.summary.items()):
                ws.cell(row=summary_row + idx + 1, column=1, value=key)
                ws.cell(row=summary_row + idx + 1, column=2, value=value)

        wb.save(filepath)

    @staticmethod
    def export_to_pdf(
        report_data: ReportData,
        template: ReportTemplate,
        filepath: str
    ) -> None:
        """
        Export report to PDF format.

        Args:
            report_data: Generated report data
            template: Template used for report
            filepath: Output file path
        """
        # Set up document
        page_size = A4 if template.page_size == "A4" else letter
        doc = SimpleDocTemplate(
            filepath,
            pagesize=page_size,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )

        # Container for flowables
        elements = []
        styles = getSampleStyleSheet()

        # Add title
        if template.show_header:
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#366092'),
                spaceAfter=12,
                alignment=TA_CENTER
            )
            title = Paragraph(template.name, title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.2*inch))

        # Prepare table data
        table_data = []

        # Headers
        headers = [col.label for col in template.columns]
        table_data.append(headers)

        # Data rows
        for row in report_data.rows:
            table_row = []
            for col in template.columns:
                value = row.get(col.key, "")
                # Format value
                if col.format_string and value:
                    try:
                        value = col.format_string.format(value)
                    except:
                        pass
                table_row.append(str(value))
            table_data.append(table_row)

        # Create table
        table = Table(table_data)

        # Table styling
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        # Apply column-specific alignments
        for col_idx, col in enumerate(template.columns):
            if col.align == "right":
                table.setStyle(TableStyle([
                    ('ALIGN', (col_idx, 1), (col_idx, -1), 'RIGHT')
                ]))
            elif col.align == "center":
                table.setStyle(TableStyle([
                    ('ALIGN', (col_idx, 1), (col_idx, -1), 'CENTER')
                ]))

        elements.append(table)

        # Add summary if available
        if report_data.summary and template.show_summary:
            elements.append(Spacer(1, 0.3*inch))

            summary_style = ParagraphStyle(
                'Summary',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.black
            )

            elements.append(Paragraph("<b>SUMMARY</b>", summary_style))
            elements.append(Spacer(1, 0.1*inch))

            for key, value in report_data.summary.items():
                text = f"<b>{key}:</b> {value}"
                elements.append(Paragraph(text, summary_style))

        # Add footer if enabled
        if template.show_footer:
            elements.append(Spacer(1, 0.3*inch))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                alignment=TA_CENTER
            )
            footer_text = f"Generated: {report_data.generated_at.strftime('%Y-%m-%d %H:%M')}"
            elements.append(Paragraph(footer_text, footer_style))

        # Build PDF
        doc.build(elements)
