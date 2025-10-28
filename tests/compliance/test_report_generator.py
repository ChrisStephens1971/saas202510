"""
Tests for Compliance Report Generator

Tests:
- General Ledger report generation
- Trial Balance report generation
- PDF export functionality
- Excel export functionality
- Multi-tenant isolation
- Edge cases (empty reports, single entry, etc.)
"""

import tempfile
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

import pytest

from qa_testing.compliance import (
    ComplianceReportGenerator,
    GeneralLedgerReport,
    TrialBalanceReport,
)
from qa_testing.generators import (
    PropertyGenerator,
    FundGenerator,
)
from qa_testing.models import LedgerEntry, Fund


class TestGeneralLedgerGeneration:
    """Test General Ledger report generation."""

    def test_generate_gl_with_entries(self):
        """Generate GL report with ledger entries."""
        # Create test data
        property1 = PropertyGenerator.create()
        fund1 = FundGenerator.create(
            tenant_id=property1.tenant_id,
            property_id=property1.id,
            name="Operating Fund",
        )
        fund2 = FundGenerator.create(
            tenant_id=property1.tenant_id,
            property_id=property1.id,
            name="Reserve Fund",
        )

        # Create ledger entries
        start_date = date(2025, 1, 1)
        end_date = date(2025, 1, 31)

        entries = [
            # Operating fund entries
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund1.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 15),
                description="Membership dues",
                amount=Decimal("1000.00"),
                is_debit=True,
                account_code="1000",
                account_name="Cash",
            ),
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund1.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 20),
                description="Landscaping expense",
                amount=Decimal("500.00"),
                is_debit=False,
                account_code="5000",
                account_name="Landscaping",
            ),
            # Reserve fund entries
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund2.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 25),
                description="Reserve contribution",
                amount=Decimal("2000.00"),
                is_debit=True,
                account_code="1000",
                account_name="Cash",
            ),
        ]

        # Generate GL report
        report = ComplianceReportGenerator.generate_general_ledger(
            tenant_id=property1.tenant_id,
            ledger_entries=entries,
            funds=[fund1, fund2],
            start_date=start_date,
            end_date=end_date,
        )

        # Verify report
        assert report.tenant_id == property1.tenant_id
        assert report.start_date == start_date
        assert report.end_date == end_date
        assert report.fund_id is None  # All funds
        assert report.fund_name is None
        assert report.entry_count == 3
        assert len(report.entries) == 3

        # Verify totals
        assert report.total_debits == Decimal("3000.00")
        assert report.total_credits == Decimal("500.00")
        assert report.closing_balance == Decimal("2500.00")

        # Verify entries are sorted by date
        assert report.entries[0].entry_date == date(2025, 1, 15)
        assert report.entries[1].entry_date == date(2025, 1, 20)
        assert report.entries[2].entry_date == date(2025, 1, 25)

        # Verify running balances
        assert report.entries[0].running_balance == Decimal("1000.00")
        assert report.entries[1].running_balance == Decimal("500.00")
        assert report.entries[2].running_balance == Decimal("2500.00")

    def test_generate_gl_single_fund(self):
        """Generate GL report for single fund."""
        property1 = PropertyGenerator.create()
        fund1 = FundGenerator.create(
            tenant_id=property1.tenant_id,
            property_id=property1.id,
            name="Operating Fund",
        )
        fund2 = FundGenerator.create(
            tenant_id=property1.tenant_id,
            property_id=property1.id,
            name="Reserve Fund",
        )

        entries = [
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund1.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 15),
                description="Dues",
                amount=Decimal("1000.00"),
                is_debit=True,
                account_code="1000",
                account_name="Cash",
            ),
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund2.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 20),
                description="Reserve",
                amount=Decimal("2000.00"),
                is_debit=True,
                account_code="1000",
                account_name="Cash",
            ),
        ]

        # Generate GL for fund1 only
        report = ComplianceReportGenerator.generate_general_ledger(
            tenant_id=property1.tenant_id,
            ledger_entries=entries,
            funds=[fund1, fund2],
            start_date=date(2025, 1, 1),
            end_date=date(2025, 1, 31),
            fund_id=fund1.id,
        )

        # Verify only fund1 entries included
        assert report.fund_id == fund1.id
        assert report.fund_name == "Operating Fund"
        assert report.entry_count == 1
        assert report.total_debits == Decimal("1000.00")

    def test_generate_gl_with_opening_balance(self):
        """Generate GL report with entries before start date."""
        property1 = PropertyGenerator.create()
        fund1 = FundGenerator.create(
            tenant_id=property1.tenant_id,
            property_id=property1.id,
            name="Operating Fund",
        )

        entries = [
            # Entry before start date (opening balance)
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund1.id,
                transaction_id=uuid4(),
                entry_date=date(2024, 12, 15),
                description="Prior balance",
                amount=Decimal("5000.00"),
                is_debit=True,
                account_code="1000",
                account_name="Cash",
            ),
            # Entry in report period
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund1.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 15),
                description="Current dues",
                amount=Decimal("1000.00"),
                is_debit=True,
                account_code="1000",
                account_name="Cash",
            ),
        ]

        report = ComplianceReportGenerator.generate_general_ledger(
            tenant_id=property1.tenant_id,
            ledger_entries=entries,
            funds=[fund1],
            start_date=date(2025, 1, 1),
            end_date=date(2025, 1, 31),
        )

        # Verify opening balance
        assert report.opening_balance == Decimal("5000.00")
        assert report.entry_count == 1  # Only entry in period
        assert report.closing_balance == Decimal("6000.00")

    def test_generate_gl_empty(self):
        """Generate GL report with no entries."""
        property1 = PropertyGenerator.create()
        fund1 = FundGenerator.create(
            tenant_id=property1.tenant_id,
            property_id=property1.id,
            name="Operating Fund",
        )

        report = ComplianceReportGenerator.generate_general_ledger(
            tenant_id=property1.tenant_id,
            ledger_entries=[],
            funds=[fund1],
            start_date=date(2025, 1, 1),
            end_date=date(2025, 1, 31),
        )

        assert report.entry_count == 0
        assert len(report.entries) == 0
        assert report.total_debits == Decimal("0.00")
        assert report.total_credits == Decimal("0.00")
        assert report.opening_balance == Decimal("0.00")
        assert report.closing_balance == Decimal("0.00")
        assert report.is_balanced is True

    def test_generate_gl_multi_tenant_isolation(self):
        """Verify GL report only includes entries for specified tenant."""
        property1 = PropertyGenerator.create()
        property2 = PropertyGenerator.create()
        fund1 = FundGenerator.create(tenant_id=property1.tenant_id, property_id=property1.id)
        fund2 = FundGenerator.create(tenant_id=property2.tenant_id, property_id=property2.id)

        entries = [
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund1.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 15),
                description="Entry 1",
                amount=Decimal("1000.00"),
                is_debit=True,
                account_code="1000",
                account_name="Cash",
            ),
            LedgerEntry(
                tenant_id=property2.tenant_id,
                property_id=property2.id,
                fund_id=fund2.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 20),
                description="Entry 2",
                amount=Decimal("2000.00"),
                is_debit=True,
                account_code="1000",
                account_name="Cash",
            ),
        ]

        # Generate report for property1
        report = ComplianceReportGenerator.generate_general_ledger(
            tenant_id=property1.tenant_id,
            ledger_entries=entries,
            funds=[fund1, fund2],
            start_date=date(2025, 1, 1),
            end_date=date(2025, 1, 31),
        )

        # Verify only property1 entries
        assert report.entry_count == 1
        assert report.total_debits == Decimal("1000.00")


class TestTrialBalanceGeneration:
    """Test Trial Balance report generation."""

    def test_generate_tb_with_accounts(self):
        """Generate TB report with multiple accounts."""
        property1 = PropertyGenerator.create()
        fund1 = FundGenerator.create(
            tenant_id=property1.tenant_id,
            property_id=property1.id,
            name="Operating Fund",
        )
        fund2 = FundGenerator.create(
            tenant_id=property1.tenant_id,
            property_id=property1.id,
            name="Reserve Fund",
        )

        entries = [
            # Operating fund
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund1.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 15),
                description="Debit",
                amount=Decimal("1000.00"),
                is_debit=True,
                account_code="1000",
                account_name="Cash",
            ),
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund1.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 20),
                description="Credit",
                amount=Decimal("500.00"),
                is_debit=False,
                account_code="5000",
                account_name="Expense",
            ),
            # Reserve fund
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund2.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 25),
                description="Debit",
                amount=Decimal("2000.00"),
                is_debit=True,
                account_code="1000",
                account_name="Cash",
            ),
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund2.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 28),
                description="Credit",
                amount=Decimal("300.00"),
                is_debit=False,
                account_code="5000",
                account_name="Expense",
            ),
        ]

        report = ComplianceReportGenerator.generate_trial_balance(
            tenant_id=property1.tenant_id,
            ledger_entries=entries,
            funds=[fund1, fund2],
            as_of_date=date(2025, 1, 31),
        )

        # Verify report
        assert report.tenant_id == property1.tenant_id
        assert report.as_of_date == date(2025, 1, 31)
        assert report.account_count == 2

        # Verify totals
        assert report.total_debits == Decimal("3000.00")
        assert report.total_credits == Decimal("800.00")
        assert report.is_balanced is False  # Not balanced: debits != credits
        assert report.balance_difference == Decimal("2200.00")

        # Verify accounts sorted by name
        assert report.accounts[0].fund_name == "Operating Fund"
        assert report.accounts[1].fund_name == "Reserve Fund"

        # Verify Operating Fund balances
        op_fund = report.accounts[0]
        assert op_fund.debit_balance == Decimal("1000.00")
        assert op_fund.credit_balance == Decimal("500.00")
        assert op_fund.net_balance == Decimal("500.00")

        # Verify Reserve Fund balances
        res_fund = report.accounts[1]
        assert res_fund.debit_balance == Decimal("2000.00")
        assert res_fund.credit_balance == Decimal("300.00")
        assert res_fund.net_balance == Decimal("1700.00")

    def test_generate_tb_empty(self):
        """Generate TB report with no entries."""
        property1 = PropertyGenerator.create()
        fund1 = FundGenerator.create(tenant_id=property1.tenant_id, property_id=property1.id)

        report = ComplianceReportGenerator.generate_trial_balance(
            tenant_id=property1.tenant_id,
            ledger_entries=[],
            funds=[fund1],
            as_of_date=date(2025, 1, 31),
        )

        assert report.account_count == 0
        assert len(report.accounts) == 0
        assert report.total_debits == Decimal("0.00")
        assert report.total_credits == Decimal("0.00")
        assert report.is_balanced is True


class TestPDFExport:
    """Test PDF export functionality."""

    def test_export_gl_pdf(self):
        """Export General Ledger report to PDF."""
        property1 = PropertyGenerator.create()
        fund1 = FundGenerator.create(
            tenant_id=property1.tenant_id,
            property_id=property1.id,
            name="Operating Fund",
        )

        entries = [
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund1.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 15),
                description="Membership dues",
                amount=Decimal("1000.00"),
                is_debit=True,
                account_code="1000",
                account_name="Cash",
            ),
        ]

        report = ComplianceReportGenerator.generate_general_ledger(
            tenant_id=property1.tenant_id,
            ledger_entries=entries,
            funds=[fund1],
            start_date=date(2025, 1, 1),
            end_date=date(2025, 1, 31),
        )

        # Export to PDF
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "general_ledger.pdf"
            ComplianceReportGenerator.export_general_ledger_pdf(report, output_path)

            # Verify file created
            assert output_path.exists()
            assert output_path.stat().st_size > 0

    def test_export_tb_pdf(self):
        """Export Trial Balance report to PDF."""
        property1 = PropertyGenerator.create()
        fund1 = FundGenerator.create(
            tenant_id=property1.tenant_id,
            property_id=property1.id,
            name="Operating Fund",
        )

        entries = [
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund1.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 15),
                description="Entry",
                amount=Decimal("1000.00"),
                is_debit=True,
                account_code="1000",
                account_name="Cash",
            ),
        ]

        report = ComplianceReportGenerator.generate_trial_balance(
            tenant_id=property1.tenant_id,
            ledger_entries=entries,
            funds=[fund1],
            as_of_date=date(2025, 1, 31),
        )

        # Export to PDF
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "trial_balance.pdf"
            ComplianceReportGenerator.export_trial_balance_pdf(report, output_path)

            # Verify file created
            assert output_path.exists()
            assert output_path.stat().st_size > 0


class TestExcelExport:
    """Test Excel export functionality."""

    def test_export_gl_excel(self):
        """Export General Ledger report to Excel."""
        property1 = PropertyGenerator.create()
        fund1 = FundGenerator.create(
            tenant_id=property1.tenant_id,
            property_id=property1.id,
            name="Operating Fund",
        )

        entries = [
            LedgerEntry(
                tenant_id=property1.tenant_id,
                property_id=property1.id,
                fund_id=fund1.id,
                transaction_id=uuid4(),
                entry_date=date(2025, 1, 15),
                description="Membership dues",
                amount=Decimal("1000.00"),
                is_debit=True,
                account_code="1000",
                account_name="Cash",
            ),
        ]

        report = ComplianceReportGenerator.generate_general_ledger(
            tenant_id=property1.tenant_id,
            ledger_entries=entries,
            funds=[fund1],
            start_date=date(2025, 1, 1),
            end_date=date(2025, 1, 31),
        )

        # Export to Excel
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "general_ledger.xlsx"
            ComplianceReportGenerator.export_general_ledger_excel(report, output_path)

            # Verify file created
            assert output_path.exists()
            assert output_path.stat().st_size > 0

            # Verify Excel format by opening
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb.active

            # Check title
            assert ws['A1'].value == "General Ledger Report"

            # Check headers
            assert ws['A10'].value == "Date"
            assert ws['B10'].value == "Fund"
            assert ws['C10'].value == "Description"
